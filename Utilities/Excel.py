import openpyxl
from MyStore.Utilities import log, constants


class Excel:
    def setexcelfile(self):
        try:
            path = constants.excel_path
            sheetname = constants.sheet_name
            workbook = openpyxl.load_workbook(path)
            self.sheet = workbook.get_sheet_by_name(sheetname)
            log.info("Excel file opened")
        except:
            log.info("Excel file NOT opened")

    def get_row_num(self, testname):

        clm = constants.col_testcase_name
        for i in range(1,self.sheet.max_row +1):
            if testname == self.sheet.cell(row=i, column=clm).value:
                return i

    def get_content(self, row, clm):
        return self.sheet.cell(row=row, column=clm).value

